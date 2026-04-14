#!/usr/bin/env python3
"""Build a canonical catalog of all known flux sites for the explorer map."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Sequence
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

import pandas as pd
from openpyxl import load_workbook

try:
    from babel import Locale
except ImportError:  # pragma: no cover - Babel is available in the current environment.
    Locale = None  # type: ignore[assignment]


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXTERNAL_DIR = Path("/Users/trevorkeenan/Climate Dropbox/Trevor Keenan/24.explorerDocs")

EMPTY_TOKENS = {"", "na", "n/a", "null", "-", ".", "--", "none"}
STRICT_COORDINATE_TOLERANCE = 0.02
MAX_COORDINATE_TOLERANCE = 0.05
AVAILABILITY_REQUEST_TIMEOUT_SECONDS = 30
MAX_NEAREST_COUNTRY_DISTANCE = 0.3
MIN_NEAREST_COUNTRY_MARGIN = 0.05
MIN_NEAREST_COUNTRY_RATIO = 1.5

AMERIFLUX_FLUXNET_AVAILABILITY_URL = "https://amfcdn.lbl.gov/api/v2/data_availability/AmeriFlux/FLUXNET/CCBY4.0"
AMERIFLUX_BASE_AVAILABILITY_URL = "https://amfcdn.lbl.gov/api/v2/data_availability/AmeriFlux/BASE-BADM/CCBY4.0"
FLUXNET2015_AVAILABILITY_URL = "https://amfcdn.lbl.gov/api/v2/data_availability/FLUXNET/FLUXNET2015/CCBY4.0"
NATURAL_EARTH_COUNTRIES_GEOJSON_URL = "https://raw.githubusercontent.com/nvkelso/natural-earth-vector/master/geojson/ne_50m_admin_0_countries.geojson"
NATURAL_EARTH_COUNTRIES_CACHE_PATH = Path.home() / ".cache" / "fluxnet-data-explorer" / "ne_50m_admin_0_countries.geojson"

COUNTRY_CODE_ALIASES = {
    "GB": "UK",
}

CANONICAL_COLUMNS = [
    "site_id",
    "site_code",
    "site_name",
    "country_code",
    "country",
    "latitude",
    "longitude",
    "source_network",
    "source_system",
    "source_files",
    "source_count",
    "in_explorer",
    "has_accessible_data",
    "known_site_only",
    "needs_manual_review",
    "review_reason",
]

MAP_COLUMNS = [
    "site_id",
    "site_code",
    "site_name",
    "country_code",
    "country",
    "latitude",
    "longitude",
    "in_explorer",
    "has_accessible_data",
    "known_site_only",
    "source_network",
    "source_category",
]

REVIEW_COLUMNS = [
    "site_id",
    "site_code",
    "site_name",
    "country_code",
    "country",
    "latitude",
    "longitude",
    "source_system",
    "source_network",
    "source_files",
    "source_count",
    "coordinate_details",
    "review_reason",
]

MULTIPLE_CANDIDATE_REVIEW_PREFIX = "multiple possible candidate merges were observed:"

SITE_ID_FIELDS = (
    "site_id",
    "siteid",
    "fluxnet_id",
    "fluxnet_site_id",
    "mysitename",
    "site_code",
    "sitecode",
    "site",
)
SITE_CODE_FIELDS = (
    "site_code",
    "sitecode",
    "code",
    "tower_code",
    "towercode",
    "station_code",
)
SITE_NAME_FIELDS = ("site_name", "name", "station_name", "site")
COUNTRY_CODE_FIELDS = ("country_code", "iso2", "iso", "cc")
COUNTRY_FIELDS = ("country", "country_name", "nation")
LATITUDE_FIELDS = ("latitude", "lat", "location_lat", "site_latitude")
LONGITUDE_FIELDS = ("longitude", "lon", "lng", "long", "location_long", "site_longitude")
NETWORK_FIELDS = (
    "source_network",
    "network",
    "regional_network",
    "region",
    "hub",
    "source",
)


@dataclass(frozen=True)
class SourceSpec:
    path: Path
    display_path: str
    source_system: str
    default_source_network: str
    precedence: int
    in_explorer: bool
    has_accessible_data: bool


@dataclass
class RawSiteRecord:
    source_spec: SourceSpec
    row_number: int
    sheet_name: str
    raw_fields: Dict[str, str]
    site_id: str
    site_id_key: str
    site_code: str
    site_code_key: str
    site_name: str
    site_name_key: str
    country_code: str
    country_code_source: str
    country: str
    latitude: float | None
    longitude: float | None
    coordinate_precision: int
    source_network: str
    meaningful_field_count: int

    @property
    def source_label(self) -> str:
        if self.sheet_name:
            return f"{self.source_spec.display_path}#{self.sheet_name}"
        return self.source_spec.display_path

    @property
    def has_coordinates(self) -> bool:
        return self.latitude is not None and self.longitude is not None

    @property
    def has_country_and_coordinates(self) -> bool:
        return self.has_coordinates and bool(self.country_code)

    @property
    def has_valid_fluxnet_site_id(self) -> bool:
        return is_fluxnet_style_site_id(self.site_id)

    def priority_key(self) -> tuple[int, int, int, int]:
        return (
            1 if self.source_spec.source_system == "explorer_accessible" else 0,
            1 if self.has_valid_fluxnet_site_id else 0,
            1 if self.has_country_and_coordinates else 0,
            self.source_spec.precedence,
        )


@dataclass
class ExplorerAccessibleTruth:
    site_id_keys: set[str] = field(default_factory=set)
    country_site_code_keys: set[tuple[str, str]] = field(default_factory=set)
    warnings: list[str] = field(default_factory=list)

    def add_site_identity(self, *, site_id: str = "", country_code: str = "", site_code: str = "") -> None:
        site_id_key = normalize_site_id_key(site_id)
        if site_id_key:
            self.site_id_keys.add(site_id_key)

        country_site_code_key = derive_country_site_code_key(
            site_id=site_id,
            country_code=country_code,
            site_code=site_code,
        )
        if country_site_code_key is not None:
            self.country_site_code_keys.add(country_site_code_key)

    def add_record(self, record: RawSiteRecord) -> None:
        self.add_site_identity(
            site_id=record.site_id,
            country_code=record.country_code,
            site_code=record.site_code,
        )


@dataclass(frozen=True)
class CountryBoundaryPart:
    country_code: str
    country_name: str
    bbox: tuple[float, float, float, float]
    rings: tuple[tuple[tuple[float, float], ...], ...]

    @property
    def bbox_area(self) -> float:
        return max(self.bbox[2] - self.bbox[0], 0.0) * max(self.bbox[3] - self.bbox[1], 0.0)


@dataclass
class CountryBoundaryLookup:
    parts: list[CountryBoundaryPart]
    warnings: list[str] = field(default_factory=list)

    @classmethod
    def from_geojson(cls, payload: dict[str, Any]) -> "CountryBoundaryLookup":
        parts: list[CountryBoundaryPart] = []
        for feature in payload.get("features", []):
            if not isinstance(feature, dict):
                continue
            properties = feature.get("properties", {}) or {}
            country_code = resolve_country_boundary_code(properties)
            country_name = normalize_country_name(
                clean_value(properties.get("NAME_EN") or properties.get("NAME") or properties.get("ADMIN")),
                country_code,
            )
            geometry = feature.get("geometry", {}) or {}
            for rings in iter_country_boundary_rings(geometry):
                bbox = boundary_bbox(rings)
                if bbox is None:
                    continue
                parts.append(
                    CountryBoundaryPart(
                        country_code=country_code,
                        country_name=country_name,
                        bbox=bbox,
                        rings=rings,
                    )
                )
        parts.sort(key=lambda part: (part.bbox_area, part.country_code, part.country_name))
        return cls(parts=parts)

    def lookup(self, latitude: float | None, longitude: float | None) -> tuple[str, str] | None:
        if latitude is None or longitude is None:
            return None
        candidate_parts = [
            part
            for part in self.parts
            if point_in_bbox(longitude, latitude, part.bbox)
        ]
        for part in candidate_parts:
            if point_in_polygon_rings(longitude, latitude, part.rings):
                return part.country_code, part.country_name
        nearest_match = self.lookup_nearest(latitude, longitude)
        if nearest_match is not None:
            return nearest_match
        return None

    def lookup_nearest(self, latitude: float | None, longitude: float | None) -> tuple[str, str] | None:
        if latitude is None or longitude is None:
            return None

        best_by_country: dict[str, tuple[float, str]] = {}
        for part in self.parts:
            bbox_distance = distance_to_bbox(longitude, latitude, part.bbox)
            if bbox_distance > MAX_NEAREST_COUNTRY_DISTANCE:
                continue
            distance = distance_to_boundary_part(longitude, latitude, part)
            if distance > MAX_NEAREST_COUNTRY_DISTANCE:
                continue
            current = best_by_country.get(part.country_code)
            if current is None or distance < current[0]:
                best_by_country[part.country_code] = (distance, part.country_name)

        ranked = sorted(best_by_country.items(), key=lambda item: (item[1][0], item[0]))
        if not ranked:
            return None

        best_country_code, (best_distance, best_country_name) = ranked[0]
        if best_distance > MAX_NEAREST_COUNTRY_DISTANCE:
            return None
        if len(ranked) == 1:
            return best_country_code, best_country_name

        second_distance = ranked[1][1][0]
        if second_distance - best_distance >= MIN_NEAREST_COUNTRY_MARGIN:
            return best_country_code, best_country_name
        if best_distance == 0 or second_distance / best_distance >= MIN_NEAREST_COUNTRY_RATIO:
            return best_country_code, best_country_name
        return None


@dataclass
class CandidateMatch:
    site_index: int
    reasons: list[str]
    rank: int


@dataclass
class SiteGroup:
    records: list[RawSiteRecord] = field(default_factory=list)
    review_reasons: set[str] = field(default_factory=set)

    def add_record(self, record: RawSiteRecord) -> None:
        self.records.append(record)

    def candidate_label(self) -> str:
        for record in self.records:
            if record.site_id:
                return record.site_id
        for record in self.records:
            if record.site_name:
                return record.site_name
        return f"group-{id(self)}"


@dataclass
class CatalogBuildResult:
    canonical_sites: list[dict[str, Any]]
    review_rows: list[dict[str, Any]]
    map_rows: list[dict[str, Any]]
    summary: dict[str, Any]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--repo-root", default=str(REPO_ROOT), help="Explorer repository root.")
    parser.add_argument(
        "--external-dir",
        default=str(DEFAULT_EXTERNAL_DIR),
        help="Directory containing additional CSV/XLS/XLSX site lists.",
    )
    parser.add_argument(
        "--output-csv",
        default="assets/all_known_flux_sites.csv",
        help="Canonical CSV output path relative to repo root.",
    )
    parser.add_argument(
        "--output-json",
        default="assets/all_known_flux_sites.json",
        help="Canonical JSON output path relative to repo root.",
    )
    parser.add_argument(
        "--output-review-csv",
        default="assets/all_known_flux_sites_review.csv",
        help="Manual-review CSV output path relative to repo root.",
    )
    parser.add_argument(
        "--output-map-json",
        default="assets/all_known_flux_sites_map.json",
        help="Map-ready JSON output path relative to repo root.",
    )
    return parser.parse_args()


def normalize_header_name(name: Any) -> str:
    value = "" if name is None else str(name).strip()
    value = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", value)
    value = value.replace("%", " percent ")
    value = re.sub(r"[^A-Za-z0-9]+", "_", value)
    value = re.sub(r"_+", "_", value)
    return value.strip("_").lower()


def clean_value(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in EMPTY_TOKENS else text


def normalize_whitespace(text: str) -> str:
    return re.sub(r"\s+", " ", clean_value(text)).strip()


def normalize_name_key(text: str) -> str:
    value = normalize_ascii(normalize_whitespace(text)).lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def normalize_ascii(text: str) -> str:
    return unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")


def is_fluxnet_style_site_id(value: str) -> bool:
    return bool(re.fullmatch(r"[A-Z]{2}-[A-Z0-9]{2,}", normalize_site_id_key(value)))


def normalize_site_id_value(value: str) -> str:
    text = normalize_whitespace(value)
    if not text:
        return ""
    text = text.replace("_", "-")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"-+", "-", text)
    if re.fullmatch(r"[A-Za-z]{2}-[A-Za-z0-9]{2,}", text):
        prefix, suffix = text.split("-", 1)
        return prefix.upper() + "-" + suffix[:1].upper() + suffix[1:]
    return text


def normalize_site_id_key(value: str) -> str:
    text = normalize_site_id_value(value)
    return text.upper()


def derive_site_code(site_id: str, explicit_site_code: str = "") -> str:
    value = normalize_whitespace(explicit_site_code)
    if value:
        return value
    normalized_site_id = normalize_site_id_value(site_id)
    if not normalized_site_id:
        return ""
    if "-" in normalized_site_id and re.fullmatch(r"[A-Za-z]{2}-[A-Za-z0-9]{2,}", normalized_site_id):
        return normalized_site_id.split("-", 1)[1]
    return normalized_site_id


def normalize_site_code_key(value: str) -> str:
    text = normalize_whitespace(value).replace("_", "-")
    text = re.sub(r"\s+", "", text)
    return text.upper()


def canonicalize_country_code(value: str) -> str:
    code = normalize_whitespace(value).upper()
    return COUNTRY_CODE_ALIASES.get(code, code)


def build_country_lookups() -> tuple[dict[str, str], dict[str, str]]:
    code_to_name: dict[str, str] = {}
    name_to_code: dict[str, str] = {}

    if Locale is not None:
        territories = Locale("en").territories
        for code, name in territories.items():
            if re.fullmatch(r"[A-Z]{2}", code):
                display = str(name)
                code_to_name[code] = display
                name_to_code[normalize_name_key(display)] = code

    aliases = {
        "usa": "US",
        "u s a": "US",
        "u s": "US",
        "united states": "US",
        "united states of america": "US",
        "uk": "UK",
        "u k": "UK",
        "united kingdom": "UK",
        "great britain": "UK",
        "russian federation": "RU",
        "people s republic of china": "CN",
        "peoples republic of china": "CN",
        "pr china": "CN",
        "south korea": "KR",
        "north korea": "KP",
        "czech republic": "CZ",
        "republic of the congo": "CG",
        "democratic republic of the congo": "CD",
        "ivory coast": "CI",
        "lao pdr": "LA",
        "taiwan": "TW",
        "palestine": "PS",
        "vatican city": "VA",
    }
    code_to_name["UK"] = "United Kingdom"
    code_to_name["GB"] = "United Kingdom"
    name_to_code.update(aliases)
    return code_to_name, name_to_code


COUNTRY_CODE_TO_NAME, COUNTRY_NAME_TO_CODE = build_country_lookups()


def normalize_country_code(
    country_code_value: str,
    country_value: str,
    site_id_value: str,
) -> tuple[str, str]:
    code = canonicalize_country_code(country_code_value)
    if re.fullmatch(r"[A-Z]{2}", code):
        return code, "explicit_code"

    country_name = normalize_name_key(country_value)
    if country_name and country_name in COUNTRY_NAME_TO_CODE:
        return canonicalize_country_code(COUNTRY_NAME_TO_CODE[country_name]), "country_name"

    site_id_key = normalize_site_id_key(site_id_value)
    if re.fullmatch(r"[A-Z]{2}-[A-Z0-9]{2,}", site_id_key):
        return canonicalize_country_code(site_id_key.split("-", 1)[0]), "site_id_prefix"

    return "", ""


def normalize_country_name(country_value: str, country_code: str) -> str:
    country = normalize_whitespace(country_value)
    if re.fullmatch(r"[A-Za-z]{2}", country):
        normalized_code = canonicalize_country_code(country)
        return COUNTRY_CODE_TO_NAME.get(normalized_code, normalized_code)
    if country:
        normalized = normalize_name_key(country)
        resolved_code = COUNTRY_NAME_TO_CODE.get(normalized, "")
        if resolved_code and resolved_code in COUNTRY_CODE_TO_NAME:
            return COUNTRY_CODE_TO_NAME[resolved_code]
        return country
    if country_code:
        normalized_code = canonicalize_country_code(country_code)
        return COUNTRY_CODE_TO_NAME.get(normalized_code, normalized_code)
    return ""


def resolve_country_boundary_code(properties: dict[str, Any]) -> str:
    for key in ("ISO_A2_EH", "ISO_A2", "WB_A2", "ADM0_ISO"):
        candidate = canonicalize_country_code(clean_value(properties.get(key)))
        if re.fullmatch(r"[A-Z]{2}", candidate):
            return candidate
    return ""


def iter_country_boundary_rings(geometry: dict[str, Any]) -> Iterator[tuple[tuple[tuple[float, float], ...], ...]]:
    geometry_type = clean_value(geometry.get("type"))
    coordinates = geometry.get("coordinates", [])
    if geometry_type == "Polygon":
        rings = normalize_boundary_rings(coordinates)
        if rings:
            yield rings
        return
    if geometry_type == "MultiPolygon":
        for polygon in coordinates:
            rings = normalize_boundary_rings(polygon)
            if rings:
                yield rings


def normalize_boundary_rings(raw_rings: Any) -> tuple[tuple[tuple[float, float], ...], ...]:
    rings: list[tuple[tuple[float, float], ...]] = []
    if not isinstance(raw_rings, list):
        return ()
    for raw_ring in raw_rings:
        ring: list[tuple[float, float]] = []
        if not isinstance(raw_ring, list):
            continue
        for point in raw_ring:
            if not isinstance(point, (list, tuple)) or len(point) < 2:
                continue
            try:
                longitude = float(point[0])
                latitude = float(point[1])
            except (TypeError, ValueError):
                continue
            ring.append((longitude, latitude))
        if len(ring) >= 4:
            rings.append(tuple(ring))
    return tuple(rings)


def boundary_bbox(rings: tuple[tuple[tuple[float, float], ...], ...]) -> tuple[float, float, float, float] | None:
    coordinates = [point for ring in rings for point in ring]
    if not coordinates:
        return None
    longitudes = [point[0] for point in coordinates]
    latitudes = [point[1] for point in coordinates]
    return min(longitudes), min(latitudes), max(longitudes), max(latitudes)


def point_in_bbox(longitude: float, latitude: float, bbox: tuple[float, float, float, float]) -> bool:
    return bbox[0] <= longitude <= bbox[2] and bbox[1] <= latitude <= bbox[3]


def point_in_polygon_rings(longitude: float, latitude: float, rings: tuple[tuple[tuple[float, float], ...], ...]) -> bool:
    if not rings:
        return False
    if not point_in_ring(longitude, latitude, rings[0]):
        return False
    for hole in rings[1:]:
        if point_in_ring(longitude, latitude, hole):
            return False
    return True


def point_in_ring(longitude: float, latitude: float, ring: tuple[tuple[float, float], ...]) -> bool:
    inside = False
    if len(ring) < 4:
        return False
    previous_longitude, previous_latitude = ring[-1]
    for current_longitude, current_latitude in ring:
        if point_on_segment(
            longitude,
            latitude,
            previous_longitude,
            previous_latitude,
            current_longitude,
            current_latitude,
        ):
            return True
        intersects = ((current_latitude > latitude) != (previous_latitude > latitude))
        if intersects:
            intersection_longitude = (
                (previous_longitude - current_longitude) * (latitude - current_latitude)
                / (previous_latitude - current_latitude)
            ) + current_longitude
            if longitude <= intersection_longitude:
                inside = not inside
        previous_longitude, previous_latitude = current_longitude, current_latitude
    return inside


def point_on_segment(
    longitude: float,
    latitude: float,
    left_longitude: float,
    left_latitude: float,
    right_longitude: float,
    right_latitude: float,
) -> bool:
    epsilon = 1e-9
    cross_product = (
        (longitude - left_longitude) * (right_latitude - left_latitude)
        - (latitude - left_latitude) * (right_longitude - left_longitude)
    )
    if abs(cross_product) > epsilon:
        return False
    min_longitude = min(left_longitude, right_longitude) - epsilon
    max_longitude = max(left_longitude, right_longitude) + epsilon
    min_latitude = min(left_latitude, right_latitude) - epsilon
    max_latitude = max(left_latitude, right_latitude) + epsilon
    return min_longitude <= longitude <= max_longitude and min_latitude <= latitude <= max_latitude


def distance_to_bbox(longitude: float, latitude: float, bbox: tuple[float, float, float, float]) -> float:
    longitude_distance = 0.0
    latitude_distance = 0.0
    if longitude < bbox[0]:
        longitude_distance = bbox[0] - longitude
    elif longitude > bbox[2]:
        longitude_distance = longitude - bbox[2]
    if latitude < bbox[1]:
        latitude_distance = bbox[1] - latitude
    elif latitude > bbox[3]:
        latitude_distance = latitude - bbox[3]
    return math.hypot(longitude_distance, latitude_distance)


def point_to_segment_distance(
    longitude: float,
    latitude: float,
    left_longitude: float,
    left_latitude: float,
    right_longitude: float,
    right_latitude: float,
) -> float:
    delta_longitude = right_longitude - left_longitude
    delta_latitude = right_latitude - left_latitude
    segment_length_squared = delta_longitude * delta_longitude + delta_latitude * delta_latitude
    if segment_length_squared == 0:
        return math.hypot(longitude - left_longitude, latitude - left_latitude)
    projection = (
        (longitude - left_longitude) * delta_longitude
        + (latitude - left_latitude) * delta_latitude
    ) / segment_length_squared
    projection = max(0.0, min(1.0, projection))
    closest_longitude = left_longitude + projection * delta_longitude
    closest_latitude = left_latitude + projection * delta_latitude
    return math.hypot(longitude - closest_longitude, latitude - closest_latitude)


def distance_to_boundary_part(longitude: float, latitude: float, part: CountryBoundaryPart) -> float:
    best_distance = float("inf")
    for ring in part.rings:
        previous_longitude, previous_latitude = ring[-1]
        for current_longitude, current_latitude in ring:
            distance = point_to_segment_distance(
                longitude,
                latitude,
                previous_longitude,
                previous_latitude,
                current_longitude,
                current_latitude,
            )
            if distance < best_distance:
                best_distance = distance
            previous_longitude, previous_latitude = current_longitude, current_latitude
    return best_distance


def parse_coordinate(value: str, minimum: float, maximum: float) -> tuple[float | None, int]:
    text = normalize_whitespace(value)
    if not text:
        return None, 0
    text = text.replace(",", "")
    try:
        numeric = float(text)
    except ValueError:
        return None, 0
    if math.isnan(numeric) or math.isinf(numeric):
        return None, 0
    if numeric < minimum or numeric > maximum:
        return None, 0
    if "." in text:
        precision = len(text.split(".", 1)[1].rstrip("0"))
    else:
        precision = 0
    return numeric, precision


def normalize_publish_years(value: Any) -> list[int]:
    if isinstance(value, (list, tuple, set)):
        raw_values = list(value)
    else:
        text = clean_value(value)
        if not text:
            return []
        raw_values = re.split(r"[\s,;|/]+", text)

    years: list[int] = []
    seen: set[int] = set()
    for raw in raw_values:
        text = clean_value(raw)
        if not re.fullmatch(r"\d{4}", text):
            continue
        year = int(text)
        if year in seen:
            continue
        seen.add(year)
        years.append(year)
    years.sort()
    return years


def read_json_from_url_or_cache(url: str, cache_path: Path) -> dict[str, Any]:
    if cache_path.exists():
        return json.loads(cache_path.read_text(encoding="utf-8"))

    request = Request(url, headers={"User-Agent": "all-known-flux-sites-builder/1.0"})
    with urlopen(request, timeout=AVAILABILITY_REQUEST_TIMEOUT_SECONDS) as response:
        payload = json.load(response)
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    cache_path.write_text(json.dumps(payload, ensure_ascii=True), encoding="utf-8")
    return payload


@lru_cache(maxsize=1)
def load_country_boundary_lookup() -> CountryBoundaryLookup:
    try:
        payload = read_json_from_url_or_cache(
            NATURAL_EARTH_COUNTRIES_GEOJSON_URL,
            NATURAL_EARTH_COUNTRIES_CACHE_PATH,
        )
    except (HTTPError, URLError, TimeoutError, OSError, json.JSONDecodeError) as exc:
        return CountryBoundaryLookup(parts=[], warnings=[f"Country boundary lookup could not be loaded: {exc}"])
    return CountryBoundaryLookup.from_geojson(payload)


def infer_country_from_coordinates(
    latitude: float | None,
    longitude: float | None,
    country_lookup: CountryBoundaryLookup | None,
) -> tuple[str, str]:
    if country_lookup is None:
        return "", ""
    match = country_lookup.lookup(latitude, longitude)
    if match is None:
        return "", ""
    return match


def format_float(value: float | None) -> str:
    if value is None:
        return ""
    text = f"{value:.6f}"
    return text.rstrip("0").rstrip(".")


def first_present(fields: dict[str, str], names: Sequence[str]) -> str:
    for name in names:
        value = fields.get(name, "")
        if value:
            return value
    return ""


def row_to_normalized_fields(row: dict[str, Any]) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for key, value in row.items():
        normalized_key = normalize_header_name(key)
        if not normalized_key:
            continue
        normalized[normalized_key] = clean_value(value)
    return normalized


def iter_csv_rows(path: Path) -> Iterator[tuple[str, int, dict[str, Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row_number, row in enumerate(reader, start=2):
            if not row:
                continue
            if not any(clean_value(value) for value in row.values()):
                continue
            yield "", row_number, dict(row)


def iter_xlsx_rows(path: Path) -> Iterator[tuple[str, int, dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            continue
        headers = [normalize_header_name(value) for value in header_row]
        if not any(headers):
            continue
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            row = {headers[index]: values[index] for index in range(len(headers)) if headers[index]}
            if not row:
                continue
            if not any(clean_value(value) for value in row.values()):
                continue
            yield sheet_name, row_number, row


def iter_xls_rows(path: Path) -> Iterator[tuple[str, int, dict[str, Any]]]:
    workbook = pd.read_excel(path, sheet_name=None, dtype=object)
    for sheet_name, frame in workbook.items():
        frame = frame.dropna(how="all")
        if frame.empty:
            continue
        headers = [normalize_header_name(value) for value in frame.columns]
        for row_offset, (_, series) in enumerate(frame.iterrows(), start=2):
            row = {
                headers[index]: ("" if pd.isna(series.iloc[index]) else series.iloc[index])
                for index in range(len(headers))
                if headers[index]
            }
            if not row:
                continue
            if not any(clean_value(value) for value in row.values()):
                continue
            yield sheet_name, row_offset, row


def iter_tabular_rows(path: Path) -> Iterator[tuple[str, int, dict[str, Any]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        yield from iter_csv_rows(path)
        return
    if suffix == ".xlsx":
        yield from iter_xlsx_rows(path)
        return
    if suffix == ".xls":
        yield from iter_xls_rows(path)
        return
    raise ValueError(f"Unsupported file type: {path}")


def build_raw_record(
    fields: dict[str, str],
    source_spec: SourceSpec,
    *,
    row_number: int,
    sheet_name: str = "",
) -> RawSiteRecord | None:
    site_id = normalize_site_id_value(first_present(fields, SITE_ID_FIELDS))
    explicit_site_code = normalize_whitespace(first_present(fields, SITE_CODE_FIELDS))
    if not site_id and is_fluxnet_style_site_id(explicit_site_code):
        site_id = normalize_site_id_value(explicit_site_code)
    site_code = derive_site_code(site_id, explicit_site_code)
    site_name = normalize_whitespace(first_present(fields, SITE_NAME_FIELDS))

    country_raw = first_present(fields, COUNTRY_FIELDS)
    country_code_raw = first_present(fields, COUNTRY_CODE_FIELDS)
    country_code, country_code_source = normalize_country_code(country_code_raw, country_raw, site_id or site_code)
    country = normalize_country_name(country_raw, country_code)

    latitude, latitude_precision = parse_coordinate(first_present(fields, LATITUDE_FIELDS), -90.0, 90.0)
    longitude, longitude_precision = parse_coordinate(first_present(fields, LONGITUDE_FIELDS), -180.0, 180.0)
    coordinate_precision = min(latitude_precision, longitude_precision)
    source_network = normalize_whitespace(first_present(fields, NETWORK_FIELDS)) or source_spec.default_source_network
    meaningful_field_count = sum(1 for value in fields.values() if clean_value(value))

    if not (site_id or site_code or site_name):
        return None

    return RawSiteRecord(
        source_spec=source_spec,
        row_number=row_number,
        sheet_name=sheet_name,
        raw_fields=fields,
        site_id=site_id,
        site_id_key=normalize_site_id_key(site_id),
        site_code=site_code,
        site_code_key=normalize_site_code_key(site_code),
        site_name=site_name,
        site_name_key=normalize_name_key(site_name),
        country_code=country_code,
        country_code_source=country_code_source,
        country=country,
        latitude=latitude,
        longitude=longitude,
        coordinate_precision=coordinate_precision,
        source_network=source_network,
        meaningful_field_count=meaningful_field_count,
    )


def relative_display_path(repo_root: Path, external_dir: Path, path: Path) -> str:
    try:
        return str(path.relative_to(repo_root)).replace("\\", "/")
    except ValueError:
        try:
            relative = path.relative_to(external_dir.parent)
            return str(relative).replace("\\", "/")
        except ValueError:
            return str(path)


def discover_repo_sources(repo_root: Path, external_dir: Path) -> list[SourceSpec]:
    source_rows = [
        ("assets/shuttle_snapshot.csv", "explorer_accessible", "FLUXNET-Shuttle", 100, True, True),
        ("assets/icos_direct_fluxnet.csv", "explorer_accessible", "ICOS", 95, True, True),
        ("assets/japanflux_direct_snapshot.csv", "explorer_accessible", "JapanFlux", 90, True, True),
        ("assets/efd_curated_sites_snapshot.csv", "explorer_accessible", "EFD", 85, True, True),
        ("assets/efd_sites_snapshot.csv", "explorer_repo_metadata", "EFD", 60, False, False),
        ("assets/ameriflux_site_info.csv", "explorer_repo_metadata", "AmeriFlux", 55, False, False),
        ("assets/siteinfo_fluxnet2015.csv", "explorer_repo_metadata", "FLUXNET2015", 50, False, False),
        ("assets/site_name_metadata.csv", "explorer_repo_metadata", "Site metadata", 40, False, False),
        ("assets/site_vegetation_metadata.csv", "explorer_repo_metadata", "Vegetation metadata", 35, False, False),
    ]
    results: list[SourceSpec] = []
    for relative_path, system, network, precedence, in_explorer, accessible in source_rows:
        path = repo_root / relative_path
        if not path.exists():
            continue
        results.append(
            SourceSpec(
                path=path,
                display_path=relative_display_path(repo_root, external_dir, path),
                source_system=system,
                default_source_network=network,
                precedence=precedence,
                in_explorer=in_explorer,
                has_accessible_data=accessible,
            )
        )
    return results


def guess_external_network(path: Path) -> str:
    name = path.name.lower()
    if "pingyu" in name or "allsites" in name:
        return "Pingyu"
    if "stoy" in name:
        return "Stoy"
    if "efdc" in name:
        return "EFD"
    if "explorersitelist" in name:
        return "Explorer export"
    if "unique_flux_sites" in name:
        return "Unique flux sites"
    if "sites_not_in_explorer" in name:
        return "Sites not in explorer"
    return "External list"


def discover_external_sources(repo_root: Path, external_dir: Path) -> list[SourceSpec]:
    results: list[SourceSpec] = []
    for path in sorted(external_dir.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in {".csv", ".xlsx", ".xls"}:
            continue
        results.append(
            SourceSpec(
                path=path,
                display_path=relative_display_path(repo_root, external_dir, path),
                source_system="external_docs",
                default_source_network=guess_external_network(path),
                precedence=20,
                in_explorer=False,
                has_accessible_data=False,
            )
        )
    return results


def load_source_records(
    source_spec: SourceSpec,
    source_file_counts: Counter[str],
) -> list[RawSiteRecord]:
    records: list[RawSiteRecord] = []
    for sheet_name, row_number, row in iter_tabular_rows(source_spec.path):
        normalized_fields = row_to_normalized_fields(row)
        record = build_raw_record(normalized_fields, source_spec, row_number=row_number, sheet_name=sheet_name)
        source_file_counts[source_spec.display_path] += 1
        if record is not None:
            records.append(record)
    return records


def coordinates_close(
    left: tuple[float, float] | None,
    right: tuple[float, float] | None,
    tolerance: float = STRICT_COORDINATE_TOLERANCE,
) -> bool:
    if left is None or right is None:
        return False
    return abs(left[0] - right[0]) <= tolerance and abs(left[1] - right[1]) <= tolerance


def record_coordinates(record: RawSiteRecord) -> tuple[float, float] | None:
    if not record.has_coordinates:
        return None
    return (record.latitude or 0.0, record.longitude or 0.0)


def haversine_distance_km(
    left: tuple[float, float],
    right: tuple[float, float],
) -> float:
    left_latitude, left_longitude = left
    right_latitude, right_longitude = right
    earth_radius_km = 6371.0
    latitude_delta = math.radians(right_latitude - left_latitude)
    longitude_delta = math.radians(right_longitude - left_longitude)
    left_radians = math.radians(left_latitude)
    right_radians = math.radians(right_latitude)
    haversine = (
        math.sin(latitude_delta / 2.0) ** 2
        + math.cos(left_radians) * math.cos(right_radians) * math.sin(longitude_delta / 2.0) ** 2
    )
    return earth_radius_km * 2.0 * math.atan2(math.sqrt(haversine), math.sqrt(1.0 - haversine))


def summarize_coordinate_details(records: Sequence[RawSiteRecord]) -> str:
    grouped_sources: dict[tuple[str, str], list[str]] = defaultdict(list)
    coordinates: list[tuple[float, float]] = []

    for record in records:
        coords = record_coordinates(record)
        if coords is None:
            continue
        coordinates.append(coords)
        grouped_sources[(format_float(coords[0]), format_float(coords[1]))].append(record.source_label)

    if not grouped_sources:
        return ""

    entries = []
    for (latitude, longitude), sources in sorted(grouped_sources.items()):
        entries.append(f"({latitude}, {longitude}) <= {', '.join(sorted(set(sources)))}")

    if len(grouped_sources) < 2:
        return " | ".join(entries)

    max_distance_km = max(
        haversine_distance_km(left, right)
        for index, left in enumerate(coordinates)
        for right in coordinates[index + 1 :]
    )
    return f"max_distance_km={max_distance_km:.2f}; " + " | ".join(entries)


def choose_group_country(
    records: Sequence[RawSiteRecord],
    site_id: str,
    latitude: float | None,
    longitude: float | None,
    country_lookup: CountryBoundaryLookup | None,
) -> tuple[str, str]:
    coordinate_country_code, _ = infer_country_from_coordinates(latitude, longitude, country_lookup)
    country_code, _ = choose_country_code(records, site_id, coordinate_country_code=coordinate_country_code)
    return country_code, coordinate_country_code


def group_summary(
    group: SiteGroup,
    country_lookup: CountryBoundaryLookup | None,
) -> dict[str, Any]:
    records = group.records
    site_id = choose_site_id(records)
    site_code = choose_site_code(records, site_id)
    site_name = choose_site_name(records)
    latitude, longitude = choose_coordinates(records)
    country_code, coordinate_country_code = choose_group_country(records, site_id, latitude, longitude, country_lookup)
    return {
        "site_id": site_id,
        "site_id_key": normalize_site_id_key(site_id),
        "site_code": site_code,
        "site_code_key": normalize_site_code_key(site_code),
        "site_name": site_name,
        "site_name_key": normalize_name_key(site_name),
        "country_code": country_code,
        "coordinate_country_code": coordinate_country_code,
        "latitude": latitude,
        "longitude": longitude,
        "has_coordinates": latitude is not None and longitude is not None,
        "has_fluxnet_site_id": is_fluxnet_style_site_id(site_id),
    }


def should_merge_into_country_prefixed_group(
    target_summary: dict[str, Any],
    source_summary: dict[str, Any],
) -> bool:
    if not target_summary["has_fluxnet_site_id"] or source_summary["has_fluxnet_site_id"]:
        return False
    if not target_summary["site_code_key"] or target_summary["site_code_key"] != source_summary["site_code_key"]:
        return False
    if not target_summary["country_code"] or target_summary["country_code"] != source_summary["country_code"]:
        return False

    if target_summary["has_coordinates"] and source_summary["has_coordinates"]:
        return coordinates_close(
            (target_summary["latitude"], target_summary["longitude"]),
            (source_summary["latitude"], source_summary["longitude"]),
        )

    if target_summary["site_name_key"] and target_summary["site_name_key"] == source_summary["site_name_key"]:
        return True

    return False


def non_resolved_review_reasons(reasons: set[str]) -> set[str]:
    return {
        reason
        for reason in reasons
        if not reason.startswith(MULTIPLE_CANDIDATE_REVIEW_PREFIX)
    }


def consolidate_country_prefixed_site_id_groups(
    groups: Sequence[SiteGroup],
    country_lookup: CountryBoundaryLookup | None,
) -> list[SiteGroup]:
    summaries = [group_summary(group, country_lookup) for group in groups]
    merge_target_by_index: dict[int, int] = {}

    for source_index, source_summary in enumerate(summaries):
        if source_summary["has_fluxnet_site_id"]:
            continue
        candidate_targets = [
            target_index
            for target_index, target_summary in enumerate(summaries)
            if target_index != source_index
            and should_merge_into_country_prefixed_group(target_summary, source_summary)
        ]
        if len(candidate_targets) == 1:
            merge_target_by_index[source_index] = candidate_targets[0]

    merged_groups: list[SiteGroup] = []
    merged_group_index_by_original: dict[int, int] = {}

    for original_index, group in enumerate(groups):
        if original_index in merge_target_by_index:
            continue
        merged_groups.append(
            SiteGroup(
                records=list(group.records),
                review_reasons=set(group.review_reasons),
            )
        )
        merged_group_index_by_original[original_index] = len(merged_groups) - 1

    for source_index, target_original_index in merge_target_by_index.items():
        target_merged_index = merged_group_index_by_original.get(target_original_index)
        if target_merged_index is None:
            continue
        merged_group = merged_groups[target_merged_index]
        merged_group.records.extend(groups[source_index].records)
        merged_group.review_reasons.update(non_resolved_review_reasons(groups[source_index].review_reasons))

    return merged_groups


def group_site_id_keys(group: SiteGroup) -> set[str]:
    return {record.site_id_key for record in group.records if record.site_id_key}


def group_site_code_keys(group: SiteGroup) -> set[str]:
    return {record.site_code_key for record in group.records if record.site_code_key}


def group_country_site_code_keys(group: SiteGroup) -> set[tuple[str, str]]:
    return {
        (record.country_code, record.site_code_key)
        for record in group.records
        if record.country_code and record.site_code_key
    }


def group_site_name_keys(group: SiteGroup) -> set[str]:
    return {record.site_name_key for record in group.records if record.site_name_key}


def group_coordinates(group: SiteGroup) -> list[tuple[float, float]]:
    return [coords for coords in (record_coordinates(record) for record in group.records) if coords is not None]


def material_coordinate_conflict(record: RawSiteRecord, group: SiteGroup) -> bool:
    if not record.has_coordinates:
        return False
    candidate = record_coordinates(record)
    existing_coordinates = group_coordinates(group)
    if not existing_coordinates:
        return False
    if any(coordinates_close(candidate, existing, MAX_COORDINATE_TOLERANCE) for existing in existing_coordinates):
        return False
    return True


def find_candidate_matches(
    record: RawSiteRecord,
    groups: Sequence[SiteGroup],
    index_by_site_id: dict[str, set[int]],
    index_by_country_site_code: dict[tuple[str, str], set[int]],
    index_by_site_name: dict[str, set[int]],
    index_by_site_code: dict[str, set[int]],
) -> tuple[list[CandidateMatch], list[int]]:
    candidate_indexes: set[int] = set()
    conflict_indexes: set[int] = set()

    if record.site_id_key:
        candidate_indexes.update(index_by_site_id.get(record.site_id_key, set()))
    if record.country_code and record.site_code_key:
        candidate_indexes.update(index_by_country_site_code.get((record.country_code, record.site_code_key), set()))
    if record.site_name_key:
        candidate_indexes.update(index_by_site_name.get(record.site_name_key, set()))
    if record.site_code_key:
        candidate_indexes.update(index_by_site_code.get(record.site_code_key, set()))

    matches: list[CandidateMatch] = []
    for site_index in sorted(candidate_indexes):
        group = groups[site_index]
        reasons: list[str] = []
        rank = 0
        if record.site_id_key and record.site_id_key in group_site_id_keys(group):
            reasons.append("exact_site_id")
            rank = max(rank, 4)
        if record.country_code and record.site_code_key and (record.country_code, record.site_code_key) in group_country_site_code_keys(group):
            if material_coordinate_conflict(record, group):
                conflict_indexes.add(site_index)
            else:
                reasons.append("country_code_plus_site_code")
                rank = max(rank, 3)
        if record.site_name_key and record.has_coordinates and record.site_name_key in group_site_name_keys(group):
            if any(coordinates_close(record_coordinates(record), coords) for coords in group_coordinates(group)):
                reasons.append("site_name_plus_near_coordinates")
                rank = max(rank, 2)
        if record.site_code_key and record.has_coordinates and record.site_code_key in group_site_code_keys(group):
            if any(coordinates_close(record_coordinates(record), coords) for coords in group_coordinates(group)):
                reasons.append("site_code_plus_near_coordinates")
                rank = max(rank, 1)
        if reasons:
            matches.append(CandidateMatch(site_index=site_index, reasons=reasons, rank=rank))
    return matches, sorted(conflict_indexes)


def update_indexes(
    site_index: int,
    group: SiteGroup,
    index_by_site_id: dict[str, set[int]],
    index_by_country_site_code: dict[tuple[str, str], set[int]],
    index_by_site_name: dict[str, set[int]],
    index_by_site_code: dict[str, set[int]],
) -> None:
    for key in group_site_id_keys(group):
        index_by_site_id[key].add(site_index)
    for key in group_country_site_code_keys(group):
        index_by_country_site_code[key].add(site_index)
    for key in group_site_name_keys(group):
        index_by_site_name[key].add(site_index)
    for key in group_site_code_keys(group):
        index_by_site_code[key].add(site_index)


def merge_records(records: Sequence[RawSiteRecord]) -> list[SiteGroup]:
    groups: list[SiteGroup] = []
    index_by_site_id: dict[str, set[int]] = defaultdict(set)
    index_by_country_site_code: dict[tuple[str, str], set[int]] = defaultdict(set)
    index_by_site_name: dict[str, set[int]] = defaultdict(set)
    index_by_site_code: dict[str, set[int]] = defaultdict(set)

    def create_group(record: RawSiteRecord, review_reason: str = "") -> None:
        group = SiteGroup(records=[record])
        if review_reason:
            group.review_reasons.add(review_reason)
        groups.append(group)
        update_indexes(len(groups) - 1, group, index_by_site_id, index_by_country_site_code, index_by_site_name, index_by_site_code)

    ordered_records = sorted(
        records,
        key=lambda record: (
            record.priority_key(),
            record.meaningful_field_count,
            record.coordinate_precision,
            record.site_id_key,
            record.site_name_key,
        ),
        reverse=True,
    )

    for record in ordered_records:
        matches, conflict_indexes = find_candidate_matches(
            record,
            groups,
            index_by_site_id,
            index_by_country_site_code,
            index_by_site_name,
            index_by_site_code,
        )

        if not matches:
            if conflict_indexes:
                conflict_labels = ", ".join(groups[index].candidate_label() for index in conflict_indexes)
                create_group(
                    record,
                    f"same country_code + site_code as existing site(s) but conflicting coordinates: {conflict_labels}",
                )
            else:
                create_group(record)
            continue

        best_rank = max(match.rank for match in matches)
        top_matches = [match for match in matches if match.rank == best_rank]

        if len(top_matches) == 1:
            chosen = top_matches[0]
            group = groups[chosen.site_index]
            group.add_record(record)
            if "exact_site_id" in chosen.reasons and material_coordinate_conflict(record, group):
                group.review_reasons.add("exact site_id match but coordinates conflict materially")
            if len(matches) > 1:
                other_labels = ", ".join(
                    groups[match.site_index].candidate_label()
                    for match in matches
                    if match.site_index != chosen.site_index
                )
                if other_labels:
                    group.review_reasons.add(f"multiple possible candidate merges were observed: {other_labels}")
            update_indexes(chosen.site_index, group, index_by_site_id, index_by_country_site_code, index_by_site_name, index_by_site_code)
            continue

        candidate_labels = ", ".join(groups[match.site_index].candidate_label() for match in top_matches)
        create_group(record, f"multiple possible candidate merges: {candidate_labels}")

    return groups


def select_best_record(
    records: Sequence[RawSiteRecord],
    predicate: callable,
    *,
    prefer_precision: bool = False,
) -> RawSiteRecord | None:
    candidates = [record for record in records if predicate(record)]
    if not candidates:
        return None
    return max(
        candidates,
        key=lambda record: (
            record.priority_key(),
            record.coordinate_precision if prefer_precision else 0,
            record.meaningful_field_count,
            1 if record.source_spec.in_explorer else 0,
        ),
    )


def choose_site_id(records: Sequence[RawSiteRecord]) -> str:
    preferred = select_best_record(records, lambda record: bool(record.site_id and record.has_valid_fluxnet_site_id))
    if preferred is not None:
        return preferred.site_id
    fallback = select_best_record(records, lambda record: bool(record.site_id))
    if fallback is not None:
        return fallback.site_id
    fallback_code = select_best_record(records, lambda record: bool(record.site_code))
    return fallback_code.site_code if fallback_code is not None else ""


def choose_site_code(records: Sequence[RawSiteRecord], site_id: str) -> str:
    preferred = select_best_record(records, lambda record: bool(record.site_code))
    if preferred is not None:
        return derive_site_code(site_id or preferred.site_id, preferred.site_code)
    return derive_site_code(site_id, "")


def derive_country_site_code_key(
    *,
    site_id: str = "",
    country_code: str = "",
    site_code: str = "",
) -> tuple[str, str] | None:
    normalized_site_id = normalize_site_id_value(site_id)
    normalized_country_code = normalize_whitespace(country_code).upper()
    normalized_site_code = normalize_site_code_key(site_code)

    if not normalized_country_code:
        normalized_country_code, _ = normalize_country_code("", "", normalized_site_id)
    if not normalized_site_code:
        normalized_site_code = normalize_site_code_key(derive_site_code(normalized_site_id, site_code))

    if normalized_country_code and normalized_site_code:
        return normalized_country_code, normalized_site_code
    return None


def choose_site_name(records: Sequence[RawSiteRecord]) -> str:
    preferred = select_best_record(records, lambda record: bool(record.site_name))
    if preferred is None:
        return ""
    candidates = [record.site_name for record in records if record.site_name]
    return max(candidates, key=lambda value: (len(value), value))


def choose_country_code(
    records: Sequence[RawSiteRecord],
    site_id: str,
    coordinate_country_code: str = "",
) -> tuple[str, str]:
    def source_rank(record: RawSiteRecord) -> int:
        confidence = {
            "explicit_code": 3,
            "country_name": 2,
            "site_id_prefix": 1,
            "": 0,
        }[record.country_code_source]
        return confidence

    candidates = [record for record in records if record.country_code]
    distinct_codes = {record.country_code for record in candidates}
    if coordinate_country_code:
        if not distinct_codes or len(distinct_codes) > 1 or coordinate_country_code not in distinct_codes:
            return coordinate_country_code, "coordinates"
    if candidates:
        best = max(
            candidates,
            key=lambda record: (
                source_rank(record),
                record.priority_key(),
                record.meaningful_field_count,
            ),
        )
        return best.country_code, best.country_code_source

    if coordinate_country_code:
        return coordinate_country_code, "coordinates"

    inferred, source = normalize_country_code("", "", site_id)
    return inferred, source


def choose_country_name(
    records: Sequence[RawSiteRecord],
    country_code: str,
    coordinate_country_name: str = "",
) -> str:
    normalized_coordinate_name = (
        normalize_country_name(coordinate_country_name, country_code)
        if coordinate_country_name
        else ""
    )
    candidates = [record.country for record in records if record.country]
    if candidates:
        best = max(candidates, key=lambda value: (len(value), value))
        normalized = normalize_country_name(best, country_code)
        if normalized:
            return normalized
    if normalized_coordinate_name:
        return normalized_coordinate_name
    return COUNTRY_CODE_TO_NAME.get(country_code, country_code)


def choose_coordinates(records: Sequence[RawSiteRecord]) -> tuple[float | None, float | None]:
    preferred = select_best_record(records, lambda record: record.has_coordinates, prefer_precision=True)
    if preferred is None:
        return None, None
    return preferred.latitude, preferred.longitude


def choose_source_network(records: Sequence[RawSiteRecord]) -> str:
    preferred = select_best_record(records, lambda record: bool(record.source_network))
    if preferred is None:
        return ""
    values = [record.source_network for record in records if record.source_network]
    return max(values, key=lambda value: (len(value), value))


def choose_source_system(records: Sequence[RawSiteRecord]) -> str:
    best = max(records, key=lambda record: (record.priority_key(), record.meaningful_field_count))
    return best.source_spec.source_system


def distinct_fluxnet_site_ids(records: Sequence[RawSiteRecord]) -> set[str]:
    return {record.site_id_key for record in records if record.site_id and record.has_valid_fluxnet_site_id}


def distinct_country_codes(records: Sequence[RawSiteRecord]) -> set[str]:
    return {record.country_code for record in records if record.country_code}


def materially_conflicting_coordinates(records: Sequence[RawSiteRecord]) -> bool:
    coordinates = [record_coordinates(record) for record in records if record.has_coordinates]
    coordinates = [coords for coords in coordinates if coords is not None]
    if len(coordinates) < 2:
        return False
    for index, left in enumerate(coordinates):
        for right in coordinates[index + 1 :]:
            if not coordinates_close(left, right, MAX_COORDINATE_TOLERANCE):
                return True
    return False


def serialize_bool(value: bool) -> str:
    return "true" if value else "false"


def serialize_csv_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return serialize_bool(value)
    if isinstance(value, float):
        return format_float(value)
    return str(value)


def group_matches_explorer_accessible_truth(
    records: Sequence[RawSiteRecord],
    accessible_truth: ExplorerAccessibleTruth,
    *,
    site_id: str,
    country_code: str,
    site_code: str,
) -> bool:
    site_id_keys = {normalize_site_id_key(site_id)}
    site_id_keys.update(record.site_id_key for record in records if record.site_id_key)
    if any(site_id_key and site_id_key in accessible_truth.site_id_keys for site_id_key in site_id_keys):
        return True

    country_site_code_keys: set[tuple[str, str]] = set()
    canonical_key = derive_country_site_code_key(
        site_id=site_id,
        country_code=country_code,
        site_code=site_code,
    )
    if canonical_key is not None:
        country_site_code_keys.add(canonical_key)

    for record in records:
        record_key = derive_country_site_code_key(
            site_id=record.site_id,
            country_code=record.country_code,
            site_code=record.site_code,
        )
        if record_key is not None:
            country_site_code_keys.add(record_key)

    return any(key in accessible_truth.country_site_code_keys for key in country_site_code_keys)


def load_accessible_truth_from_availability_sources() -> ExplorerAccessibleTruth:
    truth = ExplorerAccessibleTruth()
    availability_sources = [
        ("AmeriFlux FLUXNET", AMERIFLUX_FLUXNET_AVAILABILITY_URL),
        ("AmeriFlux BASE", AMERIFLUX_BASE_AVAILABILITY_URL),
        ("FLUXNET2015", FLUXNET2015_AVAILABILITY_URL),
    ]

    for label, url in availability_sources:
        request = Request(url, headers={"User-Agent": "all-known-flux-sites-builder/1.0"})
        try:
            with urlopen(request, timeout=AVAILABILITY_REQUEST_TIMEOUT_SECONDS) as response:
                payload = json.load(response)
        except (HTTPError, URLError, TimeoutError, json.JSONDecodeError) as exc:
            truth.warnings.append(f"{label} availability could not be loaded: {exc}")
            continue

        values = payload.get("values", []) if isinstance(payload, dict) else []
        for entry in values:
            if not isinstance(entry, dict):
                continue
            site_id = normalize_site_id_value(clean_value(entry.get("site_id")))
            publish_years = normalize_publish_years(entry.get("publish_years"))
            if not site_id or not publish_years:
                continue
            truth.add_site_identity(site_id=site_id)

    return truth


def build_canonical_sites(
    groups: Sequence[SiteGroup],
    accessible_truth: ExplorerAccessibleTruth | None = None,
    country_lookup: CountryBoundaryLookup | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    canonical_sites: list[dict[str, Any]] = []
    review_rows: list[dict[str, Any]] = []

    resolved_groups = consolidate_country_prefixed_site_id_groups(groups, country_lookup)

    for group in resolved_groups:
        records = group.records
        site_id = choose_site_id(records)
        site_code = choose_site_code(records, site_id)
        site_name = choose_site_name(records)
        latitude, longitude = choose_coordinates(records)
        coordinate_country_code, coordinate_country_name = infer_country_from_coordinates(
            latitude,
            longitude,
            country_lookup,
        )
        country_code, country_code_source = choose_country_code(
            records,
            site_id,
            coordinate_country_code=coordinate_country_code,
        )
        country = choose_country_name(
            records,
            country_code,
            coordinate_country_name=coordinate_country_name if coordinate_country_code == country_code else "",
        )
        source_network = choose_source_network(records)
        source_system = choose_source_system(records)
        source_files = sorted({record.source_label for record in records})
        coordinate_details = summarize_coordinate_details(records)
        if accessible_truth is None:
            in_explorer = any(record.source_spec.in_explorer for record in records)
            has_accessible_data = any(record.source_spec.has_accessible_data for record in records)
        else:
            has_accessible_data = group_matches_explorer_accessible_truth(
                records,
                accessible_truth,
                site_id=site_id,
                country_code=country_code,
                site_code=site_code,
            )
            in_explorer = has_accessible_data
        known_site_only = not has_accessible_data

        review_reasons = set(group.review_reasons)
        if len(distinct_fluxnet_site_ids(records)) > 1:
            review_reasons.add("conflicting FLUXNET-style site_ids contributed to the merged site")
        if materially_conflicting_coordinates(records):
            review_reasons.add("materially conflicting coordinates were present across contributing records")
        country_codes = distinct_country_codes(records)
        unresolved_country_codes = {code for code in country_codes if code != country_code}
        if len(unresolved_country_codes) > 0 and country_code_source != "coordinates":
            review_reasons.add("ambiguous country_code across contributing records")
        if not country_code and any(record.country or record.site_id for record in records):
            review_reasons.add("country_code could not be inferred confidently")
        if not (site_id or site_name or site_code):
            review_reasons.add("too little information to retain confidently")
        if country_code_source == "site_id_prefix" and not any(record.country or record.country_code for record in records):
            review_reasons.add("country_code inferred only from site_id prefix")

        review_reason = "; ".join(sorted(review_reasons))
        row = {
            "site_id": site_id,
            "site_code": site_code,
            "site_name": site_name,
            "country_code": country_code,
            "country": country,
            "latitude": latitude,
            "longitude": longitude,
            "source_network": source_network,
            "source_system": source_system,
            "source_files": "; ".join(source_files),
            "source_count": len(source_files),
            "in_explorer": in_explorer,
            "has_accessible_data": has_accessible_data,
            "known_site_only": known_site_only,
            "needs_manual_review": bool(review_reasons),
            "review_reason": review_reason,
        }
        canonical_sites.append(row)

        if review_reasons:
            review_rows.append(
                {
                    "site_id": site_id,
                    "site_code": site_code,
                    "site_name": site_name,
                    "country_code": country_code,
                    "country": country,
                    "latitude": latitude,
                    "longitude": longitude,
                    "source_system": source_system,
                    "source_network": source_network,
                    "source_files": "; ".join(source_files),
                    "source_count": len(source_files),
                    "coordinate_details": coordinate_details,
                    "review_reason": review_reason,
                }
            )

    canonical_sites.sort(key=lambda row: (row["site_id"] or "", row["site_code"] or "", row["site_name"] or ""))
    review_rows.sort(key=lambda row: (row["site_id"] or "", row["site_code"] or "", row["site_name"] or ""))
    return canonical_sites, review_rows


def build_map_rows(canonical_sites: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for site in canonical_sites:
        if site.get("latitude") is None or site.get("longitude") is None:
            continue
        rows.append(
            {
                "site_id": site["site_id"],
                "site_code": site["site_code"],
                "site_name": site["site_name"],
                "country_code": site["country_code"],
                "country": site["country"],
                "latitude": site["latitude"],
                "longitude": site["longitude"],
                "in_explorer": site["in_explorer"],
                "has_accessible_data": site["has_accessible_data"],
                "known_site_only": site["known_site_only"],
                "source_network": site["source_network"],
                "source_category": "known_site_only" if site["known_site_only"] else "accessible_data_site",
            }
        )
    rows.sort(key=lambda row: (row["site_id"] or "", row["site_code"] or "", row["site_name"] or ""))
    return rows


def write_csv(path: Path, columns: Sequence[str], rows: Sequence[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(columns))
        writer.writeheader()
        for row in rows:
            writer.writerow({column: serialize_csv_value(row.get(column)) for column in columns})


def build_json_payload(
    columns: Sequence[str],
    rows: Sequence[dict[str, Any]],
    meta: dict[str, Any],
) -> dict[str, Any]:
    return {
        "meta": meta,
        "columns": list(columns),
        "rows": [[row.get(column) for column in columns] for row in rows],
    }


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=True, separators=(",", ":")), encoding="utf-8")


def version_hash(columns: Sequence[str], rows: Sequence[dict[str, Any]]) -> str:
    payload = {"columns": list(columns), "rows": [[row.get(column) for column in columns] for row in rows]}
    return "sha256:" + hashlib.sha256(
        json.dumps(payload, ensure_ascii=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()


def build_catalog(
    repo_root: Path,
    external_dir: Path,
) -> CatalogBuildResult:
    source_specs = discover_repo_sources(repo_root, external_dir) + discover_external_sources(repo_root, external_dir)
    source_file_counts: Counter[str] = Counter()
    all_records: list[RawSiteRecord] = []
    accessible_truth = ExplorerAccessibleTruth()
    country_lookup = load_country_boundary_lookup()

    for source_spec in source_specs:
        source_records = load_source_records(source_spec, source_file_counts)
        all_records.extend(source_records)
        if source_spec.has_accessible_data:
            for record in source_records:
                accessible_truth.add_record(record)

    live_accessible_truth = load_accessible_truth_from_availability_sources()
    accessible_truth.site_id_keys.update(live_accessible_truth.site_id_keys)
    accessible_truth.country_site_code_keys.update(live_accessible_truth.country_site_code_keys)
    accessible_truth.warnings.extend(live_accessible_truth.warnings)

    groups = merge_records(all_records)
    canonical_sites, review_rows = build_canonical_sites(
        groups,
        accessible_truth=accessible_truth,
        country_lookup=country_lookup,
    )
    map_rows = build_map_rows(canonical_sites)

    summary = {
        "schema_version": 1,
        "generated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "repo_root": str(repo_root),
        "external_dir": str(external_dir),
        "files_ingested": len(source_specs),
        "raw_rows_ingested": int(sum(source_file_counts.values())),
        "final_unique_sites": len(canonical_sites),
        "records_merged": int(sum(source_file_counts.values()) - len(canonical_sites)),
        "known_site_only_sites": sum(1 for row in canonical_sites if row["known_site_only"]),
        "manual_review_sites": sum(1 for row in canonical_sites if row["needs_manual_review"]),
        "map_rows": len(map_rows),
        "source_file_counts": dict(sorted(source_file_counts.items())),
        "accessible_truth_warnings": accessible_truth.warnings,
        "country_lookup_warnings": country_lookup.warnings,
    }
    summary["version"] = version_hash(CANONICAL_COLUMNS, canonical_sites)

    return CatalogBuildResult(
        canonical_sites=canonical_sites,
        review_rows=review_rows,
        map_rows=map_rows,
        summary=summary,
    )


def print_summary(summary: dict[str, Any]) -> None:
    print(f"Files ingested: {summary['files_ingested']}")
    print(f"Raw rows ingested: {summary['raw_rows_ingested']}")
    print(f"Final unique sites: {summary['final_unique_sites']}")
    print(f"Records merged: {summary['records_merged']}")
    print(f"Known-site-only sites: {summary['known_site_only_sites']}")
    print(f"Manual-review sites: {summary['manual_review_sites']}")
    print("Counts by source file:")
    for display_path, count in summary["source_file_counts"].items():
        print(f"  {display_path}: {count}")


def main() -> None:
    args = parse_args()
    repo_root = Path(args.repo_root).resolve()
    external_dir = Path(args.external_dir).resolve()
    if not external_dir.exists():
        raise FileNotFoundError(f"External directory not found: {external_dir}")

    result = build_catalog(repo_root, external_dir)

    output_csv = (repo_root / args.output_csv).resolve()
    output_json = (repo_root / args.output_json).resolve()
    output_review_csv = (repo_root / args.output_review_csv).resolve()
    output_map_json = (repo_root / args.output_map_json).resolve()

    canonical_meta = dict(result.summary)
    canonical_payload = build_json_payload(CANONICAL_COLUMNS, result.canonical_sites, canonical_meta)
    map_meta = dict(result.summary)
    map_meta["version"] = version_hash(MAP_COLUMNS, result.map_rows)
    map_payload = build_json_payload(MAP_COLUMNS, result.map_rows, map_meta)

    write_csv(output_csv, CANONICAL_COLUMNS, result.canonical_sites)
    write_json(output_json, canonical_payload)
    write_csv(output_review_csv, REVIEW_COLUMNS, result.review_rows)
    write_json(output_map_json, map_payload)

    print(f"Wrote {output_csv}")
    print(f"Wrote {output_json}")
    print(f"Wrote {output_review_csv}")
    print(f"Wrote {output_map_json}")
    print_summary(result.summary)


if __name__ == "__main__":
    main()
